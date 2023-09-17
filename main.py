import os
import logging
import calendar
import time
from datetime import datetime
import openpyxl
import schedule
import self

logging.basicConfig(filename='HomeTimeLog.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')
# Monate
dict_months = {
    1: 'Januar',
    2: 'Februar',
    3: 'März',
    4: 'April',
    5: 'Mai',
    6: 'Juni',
    7: 'Juli',
    8: 'August',
    9: 'September',
    10: 'Oktober',
    11: 'November',
    12: 'Dezember'
}

def edit_worksheet(directory, path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    days_in_month = calendar.monthrange(datetime.now().year, datetime.now().month)[1]
    if datetime.now().weekday() < 5:
        for i in range(3, days_in_month + 3):
            cell_valueA = ws.cell(row=i, column=1).value
            cell_valueB = ws.cell(row=i, column=2).value
            cell_valueC = ws.cell(row=i, column=3).value
            if not cell_valueB and not cell_valueC and actualDay(cell_valueA):
                ws.cell(row=i, column=2, value="8:00")
                ws.cell(row=i, column=3, value="12:00")
                break
    logging.info(f'Es ist Wochenende!')
    try:
        wb.save(path)
    except PermissionError:
        logging.error(f'Konnte nicht gespeichert werden, da zurzeit jemand anders die Datei bearbeitet')
    logging.info(f'Arbeitszeiten gespeichert unter {path}.')


def actualDay(cell_valueA):
    if str(cell_valueA.day).startswith(str(datetime.today().day)):
        return True
    return False


def create_new_worksheet(directory, currentmonth, currentyear):
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Tage', 'Anfang', 'Ende', '', 'Anfang', 'Ende']
    for index, header in enumerate(headers):
        ws.cell(row=1, column=index + 1, value=header)

    days_in_month = calendar.monthrange(currentyear, currentmonth)[1]
    for i in range(3, days_in_month + 3):
        ws.cell(row=i, column=1, value=f"{datetime.now().day}, {dict_months[currentmonth]}")

    if datetime.now().weekday() < 5:
        ws.cell(row=3, column=2, value="8:00")
        ws.cell(row=3, column=3, value="12:00")

    file_name = f'Arbeitszeiten_{dict_months[currentmonth]}{currentyear}.xlsx'
    file_path = os.path.join(directory, file_name)
    try:
        wb.save(file_path)
    except PermissionError:
        logging.error(f'Konnte nicht gespeichert werden, da zurzeit jemand anders die Datei bearbeitet.')

    logging.info(f'Arbeitszeiten gespeichert unter {file_path}.')


def is_first_day_of_month():
    return datetime.now().day == 1


def check_for_new_worksheet(directory):
    currentmonth = datetime.now().month
    currentyear = datetime.now().year

    if currentmonth in dict_months:
        month = dict_months[currentmonth]
    else:
        logging.info("Ungültiger Monat.")

    filetypes = [".xltx", ".xlsx"]
    nametag = f'Arbeitszeiten_{month}{currentyear}'

    for file_type in filetypes:
        path = os.path.join(directory, f'{nametag}{file_type}')
        if os.path.exists(path):
            return path
    return None


def main():
    directory = '/Users/oliveradler/Desktop/Arbeit/2023'
    path = check_for_new_worksheet(directory)

    if path or not is_first_day_of_month():
        if not path:
            currentmonth = datetime.now().month
            currentyear = datetime.now().year
            path = os.path.join(directory, f'Arbeitszeiten_{dict_months[currentmonth]}{currentyear}.xlsx')
        edit_worksheet(directory, path)
    else:
        currentmonth = datetime.now().month
        currentyear = datetime.now().year
        create_new_worksheet(directory, currentmonth, currentyear)


if __name__ == "__main__":
    schedule.every().day.at("08:00").do(main)
    while True:
        schedule.run_pending()
        time.sleep(1000)
